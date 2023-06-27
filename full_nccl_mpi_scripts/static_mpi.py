import os
from pprint import pprint
import openpyxl


max_byte = 0
min_byte = 10000000000


def format_size(size):
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
        
    if unit_index == 0:
        return str(size)
    else:
        return f"{size:.2f}{units[unit_index]}"



def extract_info(file_name, info):
    global min_byte
    global max_byte

    with open(file_name, "r") as f:
        bytes = -1
        for line in f.readlines():

            if "nThread" in line and "minBytes" in line and "maxBytes" in line:
                ss = line.split(" ")
                bytes = ss[ss.index("minBytes")+1]
                if bytes not in info:
                    info[bytes] = {"algbw": [], "busbw": [],"time":[]}
                
            if "N/A" in line:
                ss = [x for x in line.split(' ') if x != '' and x != '\n']
                if "N/A\n" in line :
                    naStr = "N/A\n"
                else :
                    naStr = "N/A" 

                time = ss[ss.index(naStr)-3]
                abw = ss[ss.index(naStr)-2]
                bbw = ss[ss.index(naStr)-1]
                
                info[bytes]["time"].append(float(time))
                info[bytes]["algbw"].append(float(abw))
                info[bytes]["busbw"].append(float(bbw))
                bytes = int(bytes)
                min_byte = min(min_byte, bytes)
                max_byte = max(max_byte, bytes)
                bytes = -1
        
def extract_from_file(dir, file):
    info = {}
    # for op in {"all_reduce", "all_gather", "reduce_scatter", "reduce", "broadcast"}:
    for op in {"all_reduce"}:
        info[op] = {}

        for iter in range(6):
            extract_info(dir+"/"+file+"_"+op+str(iter)+".txt", info[op])

        # sort
        byte = min_byte
        while byte <= max_byte:
            if file == "nccl":
                info[op][str(byte)]["time"] = sorted(info[op][str(byte)]["time"], reverse=True)
                info[op][str(byte)]["algbw"] = sorted(info[op][str(byte)]["algbw"],reverse=False)
                info[op][str(byte)]["busbw"] = sorted(info[op][str(byte)]["busbw"],reverse=False)
                
            else:
                info[op][str(byte)]["time"] = sorted(info[op][str(byte)]["time"],reverse=False)
                info[op][str(byte)]["algbw"] = sorted(info[op][str(byte)]["algbw"],reverse=True)
                info[op][str(byte)]["busbw"] = sorted(info[op][str(byte)]["busbw"],reverse=True)

            # for type in ["time", "algbw","busbw"]:
            #         i = 0
            #         while (info[op][str(byte)][type][i] > info[op][str(byte)][type][i+1]*1.5 or info[op][str(byte)][type][i]*1.5 < info[op][str(byte)][type][i+1])and i <3:
            #             i = i+1
            #         info[op][str(byte)][type]=info[op][str(byte)][type][i:i+3]

            byte = byte*2
    
    with open("./dict.txt", "w+") as f:
        pprint(info,f)
    ##pprint(info)
    return info


def write_sheet(op,ws,type,rowOffset, colOffset,title):
    global min_byte
    global max_byte
    
    # title
    for c in range(0,6):
        ws.cell(row=rowOffset, column=c+colOffset, value=title+str(c))
    ws.cell(row=rowOffset, column=3+colOffset, value=title+"avg")

    # data
    byte = min_byte
    i = 1
    while byte <= max_byte :
        sum = 0
        for c in range(0,6):
            sum += float(op[str(byte)][type][c])
            ws.cell(row=i+rowOffset, column=c+colOffset, value=op[str(byte)][type][c])

        avg = sum/6
        ws.cell(row=i+rowOffset, column=6+colOffset, value=avg)
        byte=byte*2
        i = i+1

      

def write_formula(ws, rowOffset,col1,col2,col3):
    # title
    ws[col3+str(rowOffset)]  ="(ofccl-nccl)/nccl"

    byte = min_byte
    i = 1
    while byte <= max_byte :

        excel_formula = "=("+col2+str(i+rowOffset)+"-"+col1+str(i+rowOffset)+")/"+col1+str(i+rowOffset)
        ws[col3+str(i+rowOffset)] = excel_formula
        byte=byte*2
        i = i+1


def write_y_axi(ws, rowOffset, col):
    byte = min_byte
    i = 1
    while byte <= max_byte :
        ws.cell(row=i+rowOffset, column=col, value=format_size(byte))
        byte = byte*2
        i=i+1

    
def write_excel(info, wb):


    if "nccl" in info:
        for (key,value) in info["nccl"].items():
            ws_bw = wb.create_sheet(key+"bw")
            ws_tm = wb.create_sheet(key+"tm")
            write_y_axi(ws_bw, 2,1)
            write_sheet(value, ws_bw, "algbw", 2, 2, "nccl_algbw")
            write_sheet(value, ws_bw, "busbw", 2, 12,"nccl_busbw")
            write_y_axi(ws_tm, 2,1)
            write_sheet(value, ws_tm, "time", 2, 2, "nccl_time")

    # if "occl" in info:
    #     for (key,value) in info["occl"].items():
            
    #         if key+"bw" in wb:
    #             ws_bw = wb[key+"bw"]
    #         else:
    #             ws_bw = wb.create_sheet(key+"bw")
    #         if key+"tm" in wb:
    #             ws_tm = wb[key+"tm"]
    #         else :
    #             ws_tm = wb.create_sheet(key+"tm")

    #         #write_y_axi(ws, 2,1)
    #         write_sheet(value, ws_bw, "algbw", 2, 6, "occl_algbw")
    #         write_sheet(value, ws_bw, "busbw", 2, 16,"occl_busbw")
    #         write_sheet(value, ws_tm, "time" ,2 ,6 , "occl_time")

    # if "nccl" in info and "occl" in info:
    #     for (key,value) in info["nccl"].items():
    #         ws_bw = wb[key+"bw"]
    #         ws_tm = wb[key+"tm"]
    #         write_formula(ws_bw,2,"E","I","J")
    #         write_formula(ws_bw,2,"O","S","T")
    #         write_formula(ws_tm,2,"E","I","J")


if __name__ == "__main__":
    info = {}
    info["nccl"] = extract_from_file("./mpi_res_4hosts","nccl")
    # info["occl"] = extract_from_file("./mpi_res","occl")
    #info["nccl"] = extract_occl()
    wb = openpyxl.Workbook()
    write_excel(info, wb)
    wb.save('4hosts_full_nccl_6data.xlsx')